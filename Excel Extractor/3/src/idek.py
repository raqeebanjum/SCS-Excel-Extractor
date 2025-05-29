import pandas as pd
import json
import logging
import re
from pathlib import Path
from difflib import get_close_matches
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Define fill colors
GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Single match
RED_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")   # Multiple matches
GRAY_FILL  = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")   # No match

###############################################################################
# 1) SIZE SYNONYMS: so "one half" => "1/2"
###############################################################################
size_synonyms_map = {
    "1/4":   ["1/4", "quarter", "one quarter"],
    "3/8":   ["3/8", "three eighths"],
    "1/2":   ["1/2", "one half", "half inch"],
    "3/4":   ["3/4", "three quarter", "three fourths"],
    "1":     ["1", "one"],
    "1-1/2": ["1-1/2", "one and one half", "1.5"],
    "2":     ["2", "two"],
    "3":     ["3", "three"],
    "4":     ["4", "four"],
    "6":     ["6", "six"],
    "8":     ["8", "eight"],
    "10":    ["10", "ten"],
    "12":    ["12", "twelve"],
    "16":    ["16", "sixteen"],
    "20":    ["20", "twenty"],
    "24":    ["24", "twenty four"]
    # If you have more sizes in categories.json, they need to be added here
    # to make sure that the synonyms are recognized.
}

def normalize_text(text):
    """Normalize text while preserving fractional sizes."""
    if not isinstance(text, str):
        return ""
    text = text.lower()

    # Keep common fractional sizes intact (don't split them)
    fraction_pattern = r'\b\d{1,2}/\d{1,2}\b'
    fractions = re.findall(fraction_pattern, text)  # Extract fractions before removing other punctuation

    # Remove all punctuation except "/"
    text = re.sub(r'[^\w\s/]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()

    # Restore extracted fractions into the cleaned text
    for frac in fractions:
        text = text.replace(frac.replace("/", " "), frac)  # Preserve original fraction format

    return text

def load_buckets_from_json(json_file):
    """Load buckets from JSON file."""
    try:
        with open(json_file, 'r') as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Error loading buckets JSON file: {e}")
        return {}

def create_subgroup_mapping(buckets_dict):
    
    subgroup_to_group = {}
    subgroup_lookup = {}
    
    product_groups = buckets_dict.get('productGroups', {})
    for product_group, data in product_groups.items():
        for subgroup in data.get('subgroups', []):
            # direct map
            subgroup_to_group[subgroup] = product_group

            # normalized map
            norm_sg = normalize_text(subgroup)
            subgroup_lookup[norm_sg] = subgroup
            
            # if dash exists, also map text after dash
            if "-" in norm_sg:
                after_dash = norm_sg.split("-", 1)[1].strip()
                subgroup_lookup[after_dash] = subgroup
            
            # also map each word (if >= 4 chars)
            for word in norm_sg.split():
                if len(word) >= 4:
                    subgroup_lookup.setdefault(word, subgroup)
    
    return subgroup_to_group, subgroup_lookup

###############################################################################
# 2) BUILD LOOKUP DICTS FOR ATTRIBUTES, INCLUDING SPECIAL SIZE-SYNONYMS
###############################################################################
def build_connection_lookup(buckets_dict):
    # e.g. for "RF","FF","RTJ","SW","BW","THD"
    return {normalize_text(v): v for v in buckets_dict.get("connectionTypes", [])}

def build_flange_lookup(buckets_dict):
    # e.g. for "150#","300#","400#", etc.
    return {normalize_text(v): v for v in buckets_dict.get("flangeClasses", [])}

def build_pipe_lookup(buckets_dict):
    # e.g. for "40","40s/STD","80","80s/XH", etc.
    return {normalize_text(v): v for v in buckets_dict.get("pipeClasses", [])}

def build_size_lookup(buckets_dict):
    """
    1) For each size in the JSON (e.g. "1/2","3/4","1-1/2"), see if we have synonyms.
    2) Flatten synonyms into a dict { normalized_synonym -> canonical_size }.
    """
    final_lookup = {}
    json_sizes = buckets_dict.get("sizes", [])
    for canonical_size in json_sizes:
        # If we have synonyms, iterate through them
        if canonical_size in size_synonyms_map:
            synonyms = size_synonyms_map[canonical_size]
        else:
            # If no synonyms, just use the canonical size
            synonyms = [canonical_size]
        # Fill final_lookup
        for syn in synonyms:
            norm_syn = normalize_text(syn)
            final_lookup[norm_syn] = canonical_size
    return final_lookup

###############################################################################
# 3) Fuzzy Matching Functions
###############################################################################
def find_subgroup_matches(description, subgroup_lookup, fuzzy_cutoff=0.6):
    """
    Return ALL matches (subgroup names) for a given description:
      1) Word-based direct match
      2) If none found, fuzzy match
    """
    norm_desc = normalize_text(description)
    if not norm_desc:
        return []

    matches = set()
    # Word-based (longest to shortest)
    words = sorted(norm_desc.split(), key=len, reverse=True)
    for w in words:
        if w in subgroup_lookup:
            matches.add(subgroup_lookup[w])

    # If none found, do fuzzy
    if not matches:
        close_keys = get_close_matches(norm_desc, subgroup_lookup.keys(), n=10, cutoff=fuzzy_cutoff)
        for ck in close_keys:
            matches.add(subgroup_lookup[ck])
    return list(matches)

def find_attribute_matches(value, lookup_dict, fuzzy_cutoff=0.6):
    """
    Match an attribute value against a lookup dictionary, ensuring fractions stay intact.
    """
    norm_value = normalize_text(value)
    if not norm_value:
        return []

    # 1️⃣ First, check if the entire value is an exact match
    if norm_value in lookup_dict:
        return [lookup_dict[norm_value]]

    # 2️⃣ Next, check if any word in the value (split by spaces) is a direct match
    words = norm_value.split()
    matches = set()
    for w in words:
        if w in lookup_dict:
            matches.add(lookup_dict[w])

    # 3️⃣ Finally, do fuzzy matching **only if** no exact matches were found
    if not matches:
        close_keys = get_close_matches(norm_value, lookup_dict.keys(), n=5, cutoff=fuzzy_cutoff)
        for ck in close_keys:
            matches.add(lookup_dict[ck])

    return list(matches)

###############################################################################
# 4) PROCESS DATA IN A DATAFRAME (NO COLORING YET)
###############################################################################
def process_data_in_dataframe(df, buckets_dict):
    """
    1) Add columns to df: "Matched_Subgroup", "Matched_Product_Group", "Multiple Matches".
    2) For each row, do fuzzy matching for subgroup and attributes.
    3) Store results in the DataFrame.
    4) Return the updated DataFrame.
    """
    # Build main lookups
    subgroup_to_group, subgroup_lookup = create_subgroup_mapping(buckets_dict)
    connection_lookup = build_connection_lookup(buckets_dict)
    flange_lookup     = build_flange_lookup(buckets_dict)
    pipe_lookup       = build_pipe_lookup(buckets_dict)
    size_lookup       = build_size_lookup(buckets_dict)

    # Ensure columns exist
    if "Matched_Subgroup" not in df.columns:
        df["Matched_Subgroup"] = None
    if "Matched_Product_Group" not in df.columns:
        df["Matched_Product_Group"] = None
    if "Multiple Matches" not in df.columns:
        df["Multiple Matches"] = ""

    # Identify the "Description" column
    possible_desc_cols = [c for c in df.columns if "desc" in c.lower()]
    if possible_desc_cols:
        desc_col = possible_desc_cols[0]
    else:
        # fallback if needed
        desc_col = df.columns[0]

    # Make sure attribute columns exist
    attribute_columns = ["Connection Type 1", "Flange Class", "Pipe Class", "Size"]
    for col in attribute_columns:
        if col not in df.columns:
            df[col] = ""


    # Fuzzy adjuster
    fuzzy_cutoff = 1.0

    # Process each row
    for idx, row in df.iterrows():
        description = str(row[desc_col]) if pd.notna(row[desc_col]) else ""

        # ========== SUBGROUP ==========
        sg_matches = find_subgroup_matches(description, subgroup_lookup, fuzzy_cutoff=fuzzy_cutoff)
        if len(sg_matches) == 0:
            # Leave blank instead of "No Match"
            df.at[idx, "Matched_Subgroup"] = ""
            df.at[idx, "Matched_Product_Group"] = ""
        else:
            df.at[idx, "Matched_Subgroup"] = sg_matches[0]
            df.at[idx, "Matched_Product_Group"] = subgroup_to_group.get(sg_matches[0], "")

            # Record multiple matches in "Multiple Matches" column
            if len(sg_matches) > 1:
                multi_str = "Subgroup: " + ", ".join(sg_matches)
                existing_mm = df.at[idx, "Multiple Matches"]
                if existing_mm:
                    df.at[idx, "Multiple Matches"] = existing_mm + " | " + multi_str
                else:
                    df.at[idx, "Multiple Matches"] = multi_str

        # ========== ATTRIBUTES ==========
        for attr_col, lookup_dict in [
            ("Connection Type 1", connection_lookup),
            ("Flange Class",      flange_lookup),
            ("Pipe Class",        pipe_lookup),
            ("Size",              size_lookup),
        ]:
            original_val = str(row[attr_col]) if pd.notna(row[attr_col]) else ""
            attr_matches = find_attribute_matches(original_val, lookup_dict, fuzzy_cutoff=fuzzy_cutoff)

            if len(attr_matches) == 0:
                # No match => keep original text
                pass
            elif len(attr_matches) == 1:
                # Single
                df.at[idx, attr_col] = attr_matches[0]
            else:
                # multiple
                df.at[idx, attr_col] = attr_matches[0]
                multi_str = f"{attr_col}: " + ", ".join(attr_matches)
                existing_mm = df.at[idx, "Multiple Matches"]
                if existing_mm:
                    df.at[idx, "Multiple Matches"] = existing_mm + " | " + multi_str
                else:
                    df.at[idx, "Multiple Matches"] = multi_str

    return df

###############################################################################
# 5) APPLY COLOR HIGHLIGHTS USING OPENPYXL
###############################################################################
def apply_color_highlights(final_file):
    """
    1) Open final_file with openpyxl.
    2) For each row, apply:
       - GREEN_FILL if there's exactly 1 match
       - RED_FILL if multiple
       - GRAY_FILL if no match
    3) Save in-place.
    """
    wb = load_workbook(final_file)
    ws = wb.active

    # Identify columns by header
    header_cells = {cell.value: cell.column for cell in ws[1] if cell.value}

    match_subgroup_col = header_cells.get("Matched_Subgroup")
    match_product_col  = header_cells.get("Matched_Product_Group")
    multiple_col       = header_cells.get("Multiple Matches")

    # We'll also highlight the known attribute columns
    attribute_cols = {
        "Connection Type 1": header_cells.get("Connection Type 1"),
        "Flange Class":      header_cells.get("Flange Class"),
        "Pipe Class":        header_cells.get("Pipe Class"),
        "Size":              header_cells.get("Size"),
    }
    
    for row_idx in range(2, ws.max_row + 1):
        # --- Subgroup / Product Group highlight logic ---
        if match_subgroup_col:
            subgroup_val = ws.cell(row=row_idx, column=match_subgroup_col).value or ""
            if subgroup_val == "" or subgroup_val is None:
                ws.cell(row=row_idx, column=match_subgroup_col).fill = GRAY_FILL
                if match_product_col:
                    ws.cell(row=row_idx, column=match_product_col).fill = GRAY_FILL
            else:
                # Check if multiple
                if multiple_col:
                    mm_val = ws.cell(row=row_idx, column=multiple_col).value or ""
                    if "Subgroup:" in mm_val:
                        # multiple
                        ws.cell(row=row_idx, column=match_subgroup_col).fill = RED_FILL
                        if match_product_col:
                            # Product group might be correct for the first subgroup
                            ws.cell(row=row_idx, column=match_product_col).fill = GREEN_FILL
                        # Also fill the "Multiple Matches" cell in red
                        ws.cell(row=row_idx, column=multiple_col).fill = RED_FILL
                    else:
                        # single
                        ws.cell(row=row_idx, column=match_subgroup_col).fill = GREEN_FILL
                        if match_product_col:
                            ws.cell(row=row_idx, column=match_product_col).fill = GREEN_FILL
                else:
                    ws.cell(row=row_idx, column=match_subgroup_col).fill = GREEN_FILL
                    if match_product_col:
                        ws.cell(row=row_idx, column=match_product_col).fill = GREEN_FILL

        # --- Attributes highlight logic ---
        for attr_name, col_num in attribute_cols.items():
            if not col_num:
                continue
            cell_val = ws.cell(row=row_idx, column=col_num).value or ""
            if not cell_val or cell_val == "" or cell_val == "No Match":
                ws.cell(row=row_idx, column=col_num).fill = GRAY_FILL
            else:
                # Check if we have multiple
                if multiple_col:
                    mm_val = ws.cell(row=row_idx, column=multiple_col).value or ""
                    if f"{attr_name}:" in mm_val:
                        # multiple
                        ws.cell(row=row_idx, column=col_num).fill = RED_FILL
                        ws.cell(row=row_idx, column=multiple_col).fill = RED_FILL
                    else:
                        # single
                        ws.cell(row=row_idx, column=col_num).fill = GREEN_FILL
                else:
                    ws.cell(row=row_idx, column=col_num).fill = GREEN_FILL

    wb.save(final_file)

###############################################################################
# 6) MAIN PROCESS FLOW
###############################################################################
def process_excel(excel_file, json_file):
    """
    1) Load Excel into DF
    2) Load JSON categories
    3) Do fuzzy matching => fill columns in DF
    4) Write DF to a temp file
    5) Re-open that file => apply color fill
    6) Rename to *_matched.xlsx
    """
    try:
        df = pd.read_excel(excel_file)
        buckets_dict = load_buckets_from_json(json_file)
        if not buckets_dict:
            logger.error("Could not load buckets from JSON. Exiting.")
            return

        # 1) Fuzzy matching => DF columns updated
        df = process_data_in_dataframe(df, buckets_dict)

        # 2) Write to a temp file
        temp_file = excel_file.replace('.xlsx', '_temp_matched.xlsx')
        df.to_excel(temp_file, index=False)

        # 3) Apply color highlighting
        apply_color_highlights(temp_file)

        # 4) Rename final
        final_file = excel_file.replace('.xlsx', '_matched.xlsx')
        Path(temp_file).rename(final_file)

        logger.info(f"Processing complete. Results saved to {final_file}")

    except Exception as e:
        logger.error(f"Error processing Excel file: {e}", exc_info=True)

def main():
    excel_file = "input.xlsx"
    json_file  = "categories.json"
    process_excel(excel_file, json_file)

if __name__ == "__main__":
    main()