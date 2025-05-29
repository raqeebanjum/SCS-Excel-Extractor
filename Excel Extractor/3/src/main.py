import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fuzzywuzzy import fuzz
import json
import os

from idek import process_data_in_dataframe

os.makedirs("output", exist_ok=True)


# === Config ===
fuzzy_threshold = 90
skip_columns = {"part_number", "description", "Vendor"}
output_file = "output/Output.xlsx"

# === Color styles ===
GREEN_FILL = PatternFill(start_color="66BB66", end_color="66BB66", fill_type="solid")  # Similar (soft green)
RED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")    # Different (light green)
ORANGE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Filled from other

# === Border style ===
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC")
)

# === Load categories ===
with open("/app/src/categories.json", "r") as f:
    buckets_dict = json.load(f)

# === Load Excel files ===
mistral_raw = pd.read_excel("input/1.xlsx", dtype=str).fillna("")
deepseek_raw = pd.read_excel("input/2.xlsx", dtype=str).fillna("")



# === Run normalization/extraction logic ===
mistral_norm = process_data_in_dataframe(mistral_raw.copy(), buckets_dict)
deepseek_norm = process_data_in_dataframe(deepseek_raw.copy(), buckets_dict)

# Remove "Multiple Matches" column if it exists
for df in [mistral_norm, deepseek_norm]:
    if "Multiple Matches" in df.columns:
        df.drop(columns=["Multiple Matches"], inplace=True)

# === Use DeepSeek as output base ===
output_df = deepseek_norm.copy()
output_df.to_excel(output_file, index=False)

# === Open workbook for formatting ===
wb = load_workbook(output_file)
ws = wb.active

# === Normalize empty values like "Not specified" ===
def normalize_empty(val):
    val = val.strip().lower()
    return "" if val in {"not provided", "not specified"} else val.strip()

# === Compare and color ===
for row_idx in range(2, ws.max_row + 1):  # Data rows
    for col_idx, col_name in enumerate(output_df.columns, start=1):
        if col_name in skip_columns:
            continue

        mistral_val = normalize_empty(mistral_norm.at[row_idx - 2, col_name])
        deepseek_val = normalize_empty(deepseek_norm.at[row_idx - 2, col_name])
        cell = ws.cell(row=row_idx, column=col_idx)

        if mistral_val and deepseek_val:
            try:
                if float(mistral_val) == float(deepseek_val):
                    cell.fill = GREEN_FILL
                    continue
            except ValueError:
                pass

            similarity = fuzz.token_sort_ratio(mistral_val, deepseek_val)
            if similarity >= fuzzy_threshold:
                cell.fill = GREEN_FILL
            else:
                cell.fill = RED_FILL
        elif deepseek_val:
            cell.fill = ORANGE_FILL
        elif mistral_val:
            cell.value = mistral_val
            cell.fill = ORANGE_FILL

# === Format header row ===
header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_idx, col_name in enumerate(output_df.columns, start=1):
    clean_name = col_name.replace("_", " ").replace("-", " ").title()
    cell = ws.cell(row=1, column=col_idx)
    cell.value = clean_name
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# === Set column widths & wrap text ===
for col_idx, col_name in enumerate(output_df.columns, start=1):
    col_letter = get_column_letter(col_idx)
    max_width = 40 if "description" in col_name.lower() else 20
    ws.column_dimensions[col_letter].width = max_width

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = thin_border

# === Set uniform row height ===
uniform_height = 30
for row_idx in range(2, ws.max_row + 1):
    ws.row_dimensions[row_idx].height = uniform_height

# === Freeze header row ===
ws.freeze_panes = "A2"

# === Save final Excel ===
wb.save(output_file)
print(f"Done. Saved to {output_file}")